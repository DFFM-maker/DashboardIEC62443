#!/usr/bin/env python3
"""OT Assessment IEC 62443 - Report Generator
Assessment: OT_Assessment_Impianto_X
Date: 2026-03-19
"""
import json
import os
import csv
from datetime import datetime

# ============================================================
# ASSET INVENTORY DATA
# ============================================================
ASSETS = [
    {"id":"A-001","ip":"172.16.224.21","mac":"3C:F7:D1:80:3A:8F","vendor":"Omron Corporation",
     "device_type":"PLC","device_model":"NJ501-1300","firmware_version":"Rev 2.9","serial":"0x01c5f556",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":""},
                   {"port":4840,"proto":"tcp","service":"OPC-UA","version":""},
                   {"port":9600,"proto":"tcp","service":"FINS","version":""},
                   {"port":44818,"proto":"tcp/udp","service":"EtherNet/IP","version":""}],
     "protocols_ot":["EtherNet/IP","FINS","OPC-UA"],
     "zone":"Zone 1 - PLC Zone","criticality":"H","status":"Running (0x0030)","notes":"PLC Omron documentato"},
    {"id":"A-002","ip":"172.16.224.22","mac":"3C:F7:D1:80:4E:BB","vendor":"Omron Corporation",
     "device_type":"PLC","device_model":"NJ501-1300","firmware_version":"Rev 2.9","serial":"0x01c5fbd8",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":""},
                   {"port":4840,"proto":"tcp","service":"OPC-UA","version":""},
                   {"port":9600,"proto":"tcp","service":"FINS","version":""},
                   {"port":44818,"proto":"tcp/udp","service":"EtherNet/IP","version":""}],
     "protocols_ot":["EtherNet/IP","FINS","OPC-UA"],
     "zone":"Zone 1 - PLC Zone","criticality":"H","status":"Running+MinorFault (0x0034)","notes":"Status 0x0034: minor fault attivo"},
    {"id":"A-003","ip":"172.16.224.23","mac":"3C:F7:D1:80:3A:85","vendor":"Omron Corporation",
     "device_type":"PLC","device_model":"NJ501-1300","firmware_version":"Rev 2.9","serial":"0x01c5f551",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":""},
                   {"port":4840,"proto":"tcp","service":"OPC-UA","version":""},
                   {"port":9600,"proto":"tcp","service":"FINS","version":""},
                   {"port":44818,"proto":"tcp/udp","service":"EtherNet/IP","version":""}],
     "protocols_ot":["EtherNet/IP","FINS","OPC-UA"],
     "zone":"Zone 1 - PLC Zone","criticality":"H","status":"Running (0x0030)","notes":""},
    {"id":"A-004","ip":"172.16.224.24","mac":"3C:F7:D1:80:4E:C9","vendor":"Omron Corporation",
     "device_type":"PLC","device_model":"NJ501-1300","firmware_version":"Rev 2.9","serial":"0x01c5fbe5",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":""},
                   {"port":4840,"proto":"tcp","service":"OPC-UA","version":""},
                   {"port":9600,"proto":"tcp","service":"FINS","version":""},
                   {"port":44818,"proto":"tcp/udp","service":"EtherNet/IP","version":""}],
     "protocols_ot":["EtherNet/IP","FINS","OPC-UA"],
     "zone":"Zone 1 - PLC Zone","criticality":"H","status":"Running (0x0030)","notes":""},
    {"id":"A-005","ip":"172.16.224.111","mac":"3C:F7:D1:80:3B:DB","vendor":"Omron Corporation",
     "device_type":"PLC","device_model":"NJ501-1300","firmware_version":"Rev 2.9","serial":"0x01c5f5b8",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":""},
                   {"port":4840,"proto":"tcp","service":"OPC-UA","version":""},
                   {"port":9600,"proto":"tcp","service":"FINS","version":""},
                   {"port":44818,"proto":"tcp/udp","service":"EtherNet/IP","version":""}],
     "protocols_ot":["EtherNet/IP","FINS","OPC-UA"],
     "zone":"Zone 1 - PLC Zone","criticality":"H","status":"Running+MinorFault (0x0034)",
     "notes":"*** ASSET NON DOCUMENTATO - non in inventario originale ***"},
    {"id":"A-006","ip":"172.16.224.41","mac":"00:50:FF:14:C5:37","vendor":"Hakko Electronics",
     "device_type":"HMI","device_model":"GR-series","firmware_version":"GR-HTTPD 2.20","serial":"",
     "open_ports":[{"port":80,"proto":"tcp","service":"HTTP","version":"GR-HTTPD 2.20"}],
     "protocols_ot":["HTTP"],"zone":"Zone 2 - HMI Zone","criticality":"M","status":"Online","notes":""},
    {"id":"A-007","ip":"172.16.224.42","mac":"00:50:FF:14:CF:D6","vendor":"Hakko Electronics",
     "device_type":"HMI","device_model":"GR-series","firmware_version":"GR-HTTPD 2.20","serial":"",
     "open_ports":[{"port":80,"proto":"tcp","service":"HTTP","version":"GR-HTTPD 2.20"}],
     "protocols_ot":["HTTP"],"zone":"Zone 2 - HMI Zone","criticality":"M","status":"Online","notes":""},
    {"id":"A-008","ip":"172.16.224.43","mac":"00:50:FF:15:37:85","vendor":"Hakko Electronics",
     "device_type":"HMI","device_model":"GR-series","firmware_version":"GR-HTTPD 2.20","serial":"",
     "open_ports":[{"port":80,"proto":"tcp","service":"HTTP","version":"GR-HTTPD 2.20"}],
     "protocols_ot":["HTTP"],"zone":"Zone 2 - HMI Zone","criticality":"M","status":"Online","notes":""},
    {"id":"A-009","ip":"172.16.224.44","mac":"00:50:FF:15:37:6D","vendor":"Hakko Electronics",
     "device_type":"HMI","device_model":"GR-series","firmware_version":"GR-HTTPD 2.20","serial":"",
     "open_ports":[{"port":80,"proto":"tcp","service":"HTTP","version":"GR-HTTPD 2.20"}],
     "protocols_ot":["HTTP"],"zone":"Zone 2 - HMI Zone","criticality":"M","status":"Online","notes":""},
    {"id":"A-010","ip":"172.16.224.121","mac":"00:50:FF:15:CA:46","vendor":"Hakko Electronics",
     "device_type":"HMI","device_model":"GR-series (variante)","firmware_version":"Dropbear SSH 2022.83","serial":"",
     "open_ports":[{"port":22,"proto":"tcp","service":"SSH","version":"Dropbear 2022.83"},
                   {"port":515,"proto":"tcp","service":"LPD Printer","version":""}],
     "protocols_ot":[],"zone":"Zone 2 - HMI Zone","criticality":"M","status":"Online",
     "notes":"*** ASSET NON DOCUMENTATO - SSH+LPD inattesi su HMI ***"},
    {"id":"A-011","ip":"172.16.224.201","mac":"00:60:65:AF:ED:F4","vendor":"B&R Industrial Automation",
     "device_type":"PLC/Bilancia","device_model":"B&R Idecon Codeline","firmware_version":"N/D","serial":"ID 02795",
     "open_ports":[{"port":50000,"proto":"tcp","service":"Idecon TCP","version":""}],
     "protocols_ot":["TCP/50000 Idecon"],"zone":"Zone 1 - PLC Zone","criticality":"H","status":"Online",
     "notes":"Banner espone dati prod: 226g|codeline|ID 02795|supervisor"},
    {"id":"A-012","ip":"172.16.224.202","mac":"00:60:65:AF:EE:0A","vendor":"B&R Industrial Automation",
     "device_type":"PLC/Bilancia","device_model":"B&R Idecon Codeline","firmware_version":"N/D","serial":"ID 02792",
     "open_ports":[{"port":50000,"proto":"tcp","service":"Idecon TCP","version":""}],
     "protocols_ot":["TCP/50000 Idecon"],"zone":"Zone 1 - PLC Zone","criticality":"H","status":"Online",
     "notes":"Banner espone dati prod: ORDINE_PLC_TEST|226g|ID 02792"},
    {"id":"A-013","ip":"172.16.224.203","mac":"00:60:65:AF:EE:10","vendor":"B&R Industrial Automation",
     "device_type":"PLC/Bilancia","device_model":"B&R Idecon Codeline","firmware_version":"N/D","serial":"ID 02794",
     "open_ports":[{"port":50000,"proto":"tcp","service":"Idecon TCP","version":""}],
     "protocols_ot":["TCP/50000 Idecon"],"zone":"Zone 1 - PLC Zone","criticality":"H","status":"Online",
     "notes":"Banner espone dati prod: 900g|codeline|ID 02794"},
    {"id":"A-014","ip":"172.16.224.204","mac":"00:60:65:AF:E7:56","vendor":"B&R Industrial Automation",
     "device_type":"PLC/Bilancia","device_model":"B&R Idecon Codeline","firmware_version":"N/D","serial":"ID 02793",
     "open_ports":[{"port":50000,"proto":"tcp","service":"Idecon TCP","version":""}],
     "protocols_ot":["TCP/50000 Idecon"],"zone":"Zone 1 - PLC Zone","criticality":"H","status":"Online",
     "notes":"Banner espone dati prod: ORD1000|400g|ID 02793"},
    {"id":"A-015","ip":"172.16.224.10","mac":"14:4F:D7:CD:E5:44","vendor":"D&S Cable Industries (HK)",
     "device_type":"Workstation/Hypervisor","device_model":"Sconosciuto - VMware host","firmware_version":"N/D","serial":"",
     "open_ports":[{"port":135,"proto":"tcp","service":"MSRPC","version":""},
                   {"port":139,"proto":"tcp","service":"NetBIOS","version":""},
                   {"port":445,"proto":"tcp","service":"SMB","version":""},
                   {"port":902,"proto":"tcp","service":"VMware SOAP","version":""},
                   {"port":912,"proto":"tcp","service":"VMware Auth","version":""},
                   {"port":161,"proto":"udp","service":"SNMP","version":""}],
     "protocols_ot":[],"zone":"Zone 4 - Unclassified","criticality":"H","status":"Online",
     "notes":"*** ASSET NON DOCUMENTATO - Windows+VMware in rete OT ***"},
    {"id":"A-016","ip":"172.16.239.2","mac":"EC:EB:B8:B5:09:50","vendor":"Hewlett Packard Enterprise",
     "device_type":"Network Switch","device_model":"HPE Aruba","firmware_version":"Mocana NanoSSH 6.3","serial":"",
     "open_ports":[{"port":22,"proto":"tcp","service":"SSH","version":"Mocana NanoSSH 6.3"},
                   {"port":23,"proto":"tcp","service":"Telnet","version":""},
                   {"port":80,"proto":"tcp","service":"HTTP","version":"eHTTP v2.0"}],
     "protocols_ot":[],"zone":"Zone 3 - Infrastructure","criticality":"H","status":"Online",
     "notes":"*** TELNET ABILITATO - accesso admin in chiaro ***"},
    {"id":"A-017","ip":"172.16.239.254","mac":"00:C0:A2:09:60:C4","vendor":"Intermedium A/S (Secomea)",
     "device_type":"Remote Access Router","device_model":"Secomea SiteManager","firmware_version":"Apache httpd","serial":"",
     "open_ports":[{"port":443,"proto":"tcp","service":"HTTPS","version":"Apache"},
                   {"port":8080,"proto":"tcp","service":"HTTP Mgmt","version":""}],
     "protocols_ot":[],"zone":"Zone 3 - Remote Access","criticality":"H","status":"Online - Auth required",
     "notes":"Router accesso remoto. Porta 8080 da verificare."},
]

# ============================================================
# VULNERABILITY FINDINGS
# ============================================================
FINDINGS = [
    {"id":"F-001","device_ip":"172.16.239.2","device_model":"HPE Aruba Switch",
     "title":"Telnet abilitato su switch OT — protocollo di gestione in chiaro",
     "description":"Il network switch HPE Aruba (172.16.239.2) ha il servizio Telnet (porta 23/tcp) attivo. Telnet trasmette tutte le comunicazioni incluse le credenziali di amministrazione in chiaro sulla rete. Un attaccante con accesso al segmento di rete puo intercettare e catturare le credenziali tramite passive sniffing, compromettendo l'intera infrastruttura di rete OT.",
     "cvss_score":9.1,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N",
     "severity":"Critical","iec62443_sr":["SR 3.1","SR 4.1","SR 7.7"],"iec62443_part":"3-3",
     "evidence":"nmap 172.16.239.2: 23/tcp open telnet\nnmap -sV: 22/tcp Mocana NanoSSH 6.3, 80/tcp eHTTP v2.0 HP switch\nTelnet APERTO senza restrizioni dalla rete OT",
     "remediation":"1. Disabilitare Telnet: 'no telnet-server' sulla CLI dello switch\n2. Utilizzare esclusivamente SSH per la gestione remota\n3. Limitare accesso SSH via ACL alle sole workstation di management\n4. Configurare timeout sessione e banner di autenticazione",
     "remediation_priority":"Immediate"},
    {"id":"F-002","device_ip":"172.16.224.201, 202, 203, 204","device_model":"B&R Idecon Bilance (4 device)",
     "title":"Accesso non autenticato TCP/50000 con esposizione dati produzione in chiaro",
     "description":"I 4 dispositivi B&R/Idecon espongono il servizio sulla porta TCP 50000 senza autenticazione. La connessione accettata restituisce immediatamente dati di produzione in tempo reale: pesi misurati (226g, 900g, 400g), codici ordine (ORDINE_PLC_TEST, ORD1000), ID dispositivo (02792-02795) e username 'supervisor'. Chiunque raggiunga la rete OT puo leggere e potenzialmente scrivere dati di processo senza credenziali.",
     "cvss_score":9.8,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H",
     "severity":"Critical","iec62443_sr":["SR 2.1","SR 4.1","SR 4.2","SR 1.2"],"iec62443_part":"3-3",
     "evidence":"nc 172.16.224.201 50000: EVENT=19/03/2026 14:34:18|||226g|codeline|ID 02795|Cod. 1012|Remote connection : Local|supervisor|\nnc 172.16.224.202 50000: EVENT=19/03/2026 14:37:07|ORDINE_PLC_TEST||226g|ID 02792|supervisor|\nnc 172.16.224.203 50000: EVENT=19/03/2026 14:35:20|||900 g|ID 02794|supervisor|\nnc 172.16.224.204 50000: EVENT=19/03/2026 14:33:54|ORD1000||400 g|ID 02793|supervisor|",
     "remediation":"1. Contattare Idecon/B&R per meccanismi di autenticazione disponibili\n2. ACL switch: accesso TCP/50000 solo da IP autorizzati\n3. Abilitare autenticazione applicativa se supportata dal protocollo\n4. Considerare tunnel VPN/TLS per le comunicazioni\n5. Non esporre questi device su segmenti non fidati",
     "remediation_priority":"Immediate"},
    {"id":"F-003","device_ip":"172.16.224.21, 22, 23, 24, 111","device_model":"Omron NJ501-1300 (5 PLC)",
     "title":"Protocollo FINS esposto senza autenticazione (porta 9600/tcp)",
     "description":"Tutti e 5 i PLC Omron NJ501-1300 espongono il protocollo FINS (Factory Interface Network Service) sulla porta 9600/tcp senza autenticazione nativa. FINS consente operazioni di lettura/scrittura sulla memoria del PLC, avvio/arresto del ciclo di programma e download/upload di firmware. Due PLC (.22 e .111) mostrano status 0x0034 (minor fault) che richiede investigazione.",
     "cvss_score":8.6,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:C/C:H/I:H/A:L",
     "severity":"High","iec62443_sr":["SR 2.1","SR 1.1","SR 3.6"],"iec62443_part":"3-3",
     "evidence":"nmap -p 9600 --script omron-info 172.16.224.21-24,111: tutti i 5 PLC rispondono al protocollo FINS\nStatus .22: 0x0034 (Running + minor fault)\nStatus .111: 0x0034 (Running + minor fault)",
     "remediation":"1. Configurare IP Filter su NJ501-1300 (Network Configurator > Security Settings > IP Filter)\n2. Limitare accesso FINS alle sole workstation engineering autorizzate\n3. Investigare causa minor fault su .22 e .111\n4. Verificare disponibilita FINS authentication nel firmware corrente",
     "remediation_priority":"Immediate"},
    {"id":"F-004","device_ip":"172.16.224.21, 22, 23, 24, 111","device_model":"Omron NJ501-1300 (5 PLC)",
     "title":"EtherNet/IP esposto con device information disclosure (porta 44818)",
     "description":"Tutti e 5 i PLC Omron espongono EtherNet/IP su porta 44818 TCP/UDP. Lo scan ENIP List Identity ha restituito senza autenticazione: vendor (Omron), modello (NJ501-1300), serial number, revision (2.9) e stato operativo. EtherNet/IP standard non prevede autenticazione delle connessioni implicite.",
     "cvss_score":7.5,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:N/A:N",
     "severity":"High","iec62443_sr":["SR 2.1","SR 4.1"],"iec62443_part":"3-3",
     "evidence":"nmap enip-info: productName: NJ501-1300, vendor: Omron Corporation (47), revision: 2.9, serialNumber: 0x01c5f556, status: 0x0030 -- senza autenticazione",
     "remediation":"1. IP filtering per limitare accesso EtherNet/IP\n2. Verificare CIP Security (estensione ODVA per autenticazione/cifratura)\n3. Disabilitare risposta ENIP List Identity a host non autorizzati\n4. Implementare firewall industriale tra HMI Zone e PLC Zone",
     "remediation_priority":"Short-term"},
    {"id":"F-005","device_ip":"172.16.224.41, 42, 43, 44","device_model":"Hakko GR-series HMI (4 device)",
     "title":"Interfaccia web HMI accessibile via HTTP non cifrato (porta 80)",
     "description":"Le 4 HMI Hakko espongono l'interfaccia web di supervisione unicamente via HTTP senza HTTPS. Credenziali di accesso e dati operativi vengono trasmessi in chiaro sulla rete OT. Server identificato: GR-HTTPD Server/2.20 (HTTP/1.0).",
     "cvss_score":6.5,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:N/A:N",
     "severity":"High","iec62443_sr":["SR 4.1","SR 3.1"],"iec62443_part":"3-3",
     "evidence":"curl http://172.16.224.41/: HTTP/1.0 404 Not Found, Server: GR-HTTPD Server/2.20 (nessun redirect HTTPS)",
     "remediation":"1. Verificare se il firmware Hakko GR supporta HTTPS e aggiornare\n2. Se HTTPS non supportato: implementare reverse proxy TLS\n3. ACL switch: limitare accesso HTTP alle sole workstation operatore\n4. Valutare VLAN dedicata HMI con accesso controllato",
     "remediation_priority":"Short-term"},
    {"id":"F-006","device_ip":"172.16.224.10","device_model":"Host Windows/VMware (non documentato)",
     "title":"Host Windows con SMB e servizi VMware nel segmento OT industriale",
     "description":"Il dispositivo 172.16.224.10 espone porte tipiche Windows (135 MSRPC, 139 NetBIOS, 445 SMB) e VMware (902, 912). La presenza di SMB in una rete OT industriale e un vettore primario per ransomware e lateral movement. Il dispositivo non era documentato nell'inventario originale e la sua presenza non e giustificata.",
     "cvss_score":7.8,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N",
     "severity":"High","iec62443_sr":["SR 1.1","SR 2.1","SR 7.7"],"iec62443_part":"3-3",
     "evidence":"nmap 172.16.224.10: 135/tcp msrpc, 139/tcp netbios, 445/tcp microsoft-ds, 902/tcp VMware, 912/tcp VMware, 161/udp snmp open|filtered\nMAC: 14:4F:D7:CD:E5:44 (D&S Cable Industries HK)",
     "remediation":"1. Identificare immediatamente il device e il suo ruolo\n2. Se non necessario nel segmento OT: rimuovere o isolare in VLAN separata\n3. Disabilitare SMBv1, applicare patch Windows correnti\n4. Firewall: bloccare porte 135,139,445 dai device OT\n5. Se VMware ESXi: aggiornare e limitare accesso management\n6. Documentare nell'inventario ufficiale",
     "remediation_priority":"Immediate"},
    {"id":"F-007","device_ip":"172.16.224.111, 172.16.224.121","device_model":"Omron NJ501-1300 + Hakko GR",
     "title":"Asset non documentati nell'inventario di sicurezza OT",
     "description":"Due asset attivi nella rete OT non erano presenti nell'inventario originale: (1) 172.16.224.111 - PLC Omron NJ501-1300 rev 2.9 con minor fault (0x0034), identico agli altri 4 PLC. (2) 172.16.224.121 - HMI Hakko che espone SSH Dropbear 2022.83 e LPD printer (515), senza HTTP a differenza degli altri HMI. Asset non documentati sfuggono ai processi di patch management e hardening.",
     "cvss_score":5.3,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:L/I:L/A:N",
     "severity":"Medium","iec62443_sr":["SR 1.1","SR 2.4"],"iec62443_part":"3-3",
     "evidence":"nmap discovery: trovati .111 (MAC Omron 3C:F7:D1:80:3B:DB) e .121 (MAC Hakko 00:50:FF:15:CA:46)\nenip-info .111: NJ501-1300 rev 2.9, status 0x0034\nnmap .121: 22/tcp Dropbear sshd 2022.83, 515/tcp printer",
     "remediation":"1. Documentare entrambi i device nell'inventario ufficiale\n2. Investigare e risolvere il minor fault su .111\n3. Verificare su .121 se SSH/LPD sono configurazioni intenzionali\n4. Disabilitare servizi non necessari su .121\n5. Implementare discovery automatica periodica (es. nmap settimanale + alert nuovi host)",
     "remediation_priority":"Short-term"},
    {"id":"F-008","device_ip":"172.16.224.21, 22, 23, 24, 111","device_model":"Omron NJ501-1300 (5 PLC)",
     "title":"OPC-UA esposto — security mode non verificato (porta 4840)",
     "description":"Tutti e 5 i PLC espongono OPC-UA sulla porta 4840. La configurazione di default di OPC-UA potrebbe non richiedere autenticazione (security mode 'None'). Lo script NSE opcua-info non era disponibile per verifica approfondita. Raccomandato test manuale con UA Expert per verificare le security policy configurate.",
     "cvss_score":6.5,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:N/A:N",
     "severity":"Medium","iec62443_sr":["SR 2.1","SR 4.1"],"iec62443_part":"3-3",
     "evidence":"nmap: 4840/tcp open opcua-tcp su tutti i 5 PLC. Script opcua-info non disponibile su Kali.",
     "remediation":"1. Verificare con UA Expert il security mode configurato\n2. Configurare: Security Mode = Sign&Encrypt, Policy = Basic256Sha256\n3. Abilitare autenticazione con certificati X.509 o username/password\n4. Disabilitare security mode 'None'\n5. Configurare whitelist degli endpoint OPC-UA autorizzati",
     "remediation_priority":"Short-term"},
    {"id":"F-009","device_ip":"172.16.239.254","device_model":"Secomea SiteManager",
     "title":"Porta HTTP 8080 aperta su router accesso remoto",
     "description":"Il router Secomea SiteManager espone la porta 8080 accessibile dalla rete OT interna. Se questa porta espone un'interfaccia non autenticata o con credenziali deboli potrebbe consentire la compromissione del gateway di accesso remoto, con impatto su tutta la catena di accesso esterno.",
     "cvss_score":5.0,"cvss_vector":"CVSS:3.1/AV:N/AC:H/PR:N/UI:N/S:U/C:L/I:L/A:L",
     "severity":"Medium","iec62443_sr":["SR 1.1","SR 2.1"],"iec62443_part":"3-3",
     "evidence":"nmap: 443/tcp https + 8080/tcp http-proxy su 172.16.239.254\ncurl 443/tcp: 401 Authorization Required (HTTPS con auth OK)\ncurl 8080/tcp: non testato per prudenza",
     "remediation":"1. Verificare utilizzo porta 8080 (management? API? redirect?)\n2. Se non necessaria: disabilitarla dalla configurazione Secomea\n3. Se necessaria: assicurarsi richieda autenticazione forte\n4. Limitare accesso porta 8080 tramite ACL switch",
     "remediation_priority":"Short-term"},
    {"id":"F-010","device_ip":"172.16.224.121","device_model":"Hakko GR (HMI .121)",
     "title":"Servizi inattesi su HMI: SSH Dropbear e LPD printer (porta 515)",
     "description":"L'HMI Hakko 172.16.224.121 presenta un profilo di porte anomalo rispetto agli altri HMI: espone SSH (Dropbear sshd 2022.83) e LPD printer (515) invece di HTTP. Dropbear e una implementazione SSH lightweight per sistemi embedded Linux. La porta 515 LPD e obsoleta e storicamente vulnerabile. Il profilo diverso suggerisce una configurazione non standard o modificata.",
     "cvss_score":4.3,"cvss_vector":"CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:L/I:N/A:N",
     "severity":"Medium","iec62443_sr":["SR 1.1","SR 2.4"],"iec62443_part":"3-3",
     "evidence":"nmap .121: 22/tcp open Dropbear sshd 2022.83, 515/tcp open printer\n(altri HMI .41-.44: solo 80/tcp HTTP GR-HTTPD 2.20)",
     "remediation":"1. Investigare se SSH e LPD sono configurazioni intenzionali\n2. Se non necessari: disabilitarli dalla configurazione HMI\n3. Verificare integrita del sistema (possibile installazione non standard)\n4. Documentare le differenze rispetto agli altri HMI",
     "remediation_priority":"Short-term"},
]

# ============================================================
# IEC 62443-3-3 SR COMPLIANCE TABLE
# ============================================================
SR_TABLE = [
    {"sr":"SR 1.1","title":"Human user identification and authentication","status":"Non Conforme","note":"HMI HTTP senza auth, B&R TCP/50000 senza auth"},
    {"sr":"SR 1.2","title":"Software process and device identification","status":"Non Conforme","note":"B&R espone dati identificazione senza autenticazione"},
    {"sr":"SR 1.3","title":"Account management","status":"Parziale","note":"Credenziali supervisor esposte (B&R)"},
    {"sr":"SR 1.4","title":"Identifier management","status":"Parziale","note":"IP statici assegnati ma 3 asset non documentati"},
    {"sr":"SR 1.5","title":"Authenticator management","status":"N/A","note":"Non verificabile senza accesso alle configurazioni"},
    {"sr":"SR 1.6","title":"Wireless access management","status":"N/A","note":"Nessun device wireless identificato"},
    {"sr":"SR 1.7","title":"Strength of password-based authentication","status":"Non Conforme","note":"Telnet espone password in chiaro"},
    {"sr":"SR 1.8","title":"Public key infrastructure certificates","status":"Parziale","note":"HTTPS su Omron e Secomea, ma HTTP su HMI e switch"},
    {"sr":"SR 1.9","title":"Strength of public key authentication","status":"Parziale","note":"SSH disponibile su switch (.239.2) ma Telnet prevalente"},
    {"sr":"SR 1.10","title":"Authenticator feedback","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 1.11","title":"Unsuccessful login attempts","status":"N/A","note":"Non verificabile senza accesso diretto"},
    {"sr":"SR 1.12","title":"System use notification","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 1.13","title":"Access via untrusted networks","status":"Non Conforme","note":"Rete OT flat, nessuna segmentazione firewall verificata"},
    {"sr":"SR 2.1","title":"Authorization enforcement","status":"Non Conforme","note":"B&R TCP/50000 senza auth, FINS senza auth, EtherNet/IP senza auth"},
    {"sr":"SR 2.2","title":"Wireless use control","status":"N/A","note":"Nessun device wireless identificato"},
    {"sr":"SR 2.3","title":"Use of portable and mobile devices","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 2.4","title":"Mobile code","status":"Non Conforme","note":"Asset non documentati - gap inventario"},
    {"sr":"SR 2.5","title":"Session lock","status":"N/A","note":"Non verificabile passivamente"},
    {"sr":"SR 2.6","title":"Remote session termination","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 2.7","title":"Concurrent session control","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 2.8","title":"Auditable events","status":"Parziale","note":"B&R logga eventi (visibili nel banner), ma senza cifratura"},
    {"sr":"SR 2.9","title":"Audit storage capacity","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 2.10","title":"Response to audit processing failures","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 2.11","title":"Timestamps","status":"Conforme","note":"B&R include timestamp negli eventi"},
    {"sr":"SR 2.12","title":"Non-repudiation","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 3.1","title":"Communication integrity","status":"Non Conforme","note":"Telnet su switch, HTTP su HMI, TCP/50000 in chiaro"},
    {"sr":"SR 3.2","title":"Malicious code protection","status":"N/A","note":"Non verificabile passivamente"},
    {"sr":"SR 3.3","title":"Security functionality verification","status":"N/A","note":"Non eseguito in questa fase"},
    {"sr":"SR 3.4","title":"Software and information integrity","status":"N/A","note":"Non verificabile passivamente"},
    {"sr":"SR 3.5","title":"Input validation","status":"N/A","note":"Non testato"},
    {"sr":"SR 3.6","title":"Deterministic output","status":"Parziale","note":"FINS esposto puo alterare output PLC"},
    {"sr":"SR 3.7","title":"Error handling","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 3.8","title":"Session integrity","status":"Non Conforme","note":"Telnet non prevede protezione integrita sessione"},
    {"sr":"SR 3.9","title":"Protection of audit information","status":"Non Conforme","note":"Log B&R esposti in chiaro su TCP/50000"},
    {"sr":"SR 4.1","title":"Information confidentiality","status":"Non Conforme","note":"Dati produzione B&R in chiaro, credenziali Telnet in chiaro, HTTP HMI in chiaro"},
    {"sr":"SR 4.2","title":"Use of cryptography","status":"Parziale","note":"HTTPS su Omron/Secomea, ma Telnet/HTTP/TCP50000 senza cifratura"},
    {"sr":"SR 4.3","title":"Use of cryptography - key management","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 5.1","title":"Network segmentation","status":"Non Conforme","note":"Rete OT flat, nessuna segmentazione Zone/Conduit conforme IEC 62443-3-2 verificata"},
    {"sr":"SR 5.2","title":"Zone boundary protection","status":"Non Conforme","note":"Nessun firewall industriale identificato tra zone"},
    {"sr":"SR 5.3","title":"General purpose person-to-person communication restrictions","status":"N/A","note":"Non applicabile"},
    {"sr":"SR 5.4","title":"Application partitioning","status":"Non Conforme","note":"Windows/VMware (.10) in stessa zona dei PLC"},
    {"sr":"SR 6.1","title":"Audit log accessibility","status":"Parziale","note":"Log B&R accessibili (ma senza protezione)"},
    {"sr":"SR 6.2","title":"Continuous monitoring","status":"Non Conforme","note":"Nessun sistema di monitoring continuo identificato"},
    {"sr":"SR 7.1","title":"Denial of service protection","status":"N/A","note":"Non testato (fuori scope)"},
    {"sr":"SR 7.2","title":"Resource management","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 7.3","title":"Control system backup","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 7.4","title":"Control system recovery and reconstitution","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 7.5","title":"Emergency power","status":"N/A","note":"Non verificabile"},
    {"sr":"SR 7.6","title":"Network and security configuration settings","status":"Non Conforme","note":"Telnet abilitato, default config rilevata"},
    {"sr":"SR 7.7","title":"Least functionality","status":"Non Conforme","note":"Telnet, SMB, LPD, VMware ports non necessari in rete OT"},
    {"sr":"SR 7.8","title":"Control system component inventory","status":"Non Conforme","note":"3 asset non documentati trovati (.10, .111, .121)"},
]

# ============================================================
# GENERATE CSV FILES
# ============================================================
def gen_csv(base_path):
    # assets.csv
    with open(os.path.join(base_path, 'assets.csv'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'IP', 'MAC', 'Vendor', 'Tipo', 'Modello', 'Firmware/Version',
                         'Serial', 'Porte Aperte', 'Protocolli OT', 'Security Zone',
                         'Criticita', 'Status', 'Note', 'Data Rilevamento'])
        for a in ASSETS:
            ports_str = '; '.join(f"{p['port']}/{p['proto']}({p['service']})" for p in a['open_ports'])
            writer.writerow([
                a['id'], a['ip'], a['mac'], a['vendor'], a['device_type'], a['device_model'],
                a['firmware_version'], a.get('serial',''), ports_str,
                ', '.join(a['protocols_ot']), a['zone'], a['criticality'],
                a['status'], a['notes'], '2026-03-19'
            ])
    print("assets.csv OK")

    # findings.csv
    with open(os.path.join(base_path, 'findings.csv'), 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['ID', 'IP Device', 'Modello', 'Titolo', 'Severita', 'CVSS Score',
                         'CVSS Vector', 'IEC 62443 SR', 'IEC 62443 Part', 'Evidence',
                         'Remediation', 'Priorita', 'Descrizione'])
        for fin in FINDINGS:
            writer.writerow([
                fin['id'], fin['device_ip'], fin['device_model'], fin['title'],
                fin['severity'], fin['cvss_score'], fin['cvss_vector'],
                '; '.join(fin['iec62443_sr']), fin['iec62443_part'],
                fin['evidence'], fin['remediation'], fin['remediation_priority'],
                fin['description']
            ])
    print("findings.csv OK")

# ============================================================
# GENERATE EXCEL FILE
# ============================================================
def gen_excel(base_path):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl non disponibile, skip Excel")
        return

    wb = openpyxl.Workbook()

    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
    SEVERITY_COLORS = {
        "Critical": "FF0000", "High": "FF6600",
        "Medium": "FFCC00", "Low": "00AA00", "Info": "0070C0"
    }
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row):
        for cell in ws[row]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def autofit(ws, min_w=10, max_w=50):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    v = str(cell.value) if cell.value else ''
                    max_len = max(max_len, min(len(v), 80))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

    # ---- Sheet 1: Asset Inventory ----
    ws1 = wb.active
    ws1.title = "Asset Inventory"
    headers1 = ['ID Asset','IP','MAC Address','Vendor','Tipo Device','Modello',
                 'Firmware/Version','Serial','Porte Aperte','Protocolli OT',
                 'Security Zone','Criticita','Status','Note','Data Rilevamento']
    ws1.append(headers1)
    style_header(ws1, 1)
    ws1.freeze_panes = 'A2'
    for i, a in enumerate(ASSETS):
        ports_str = ', '.join(f"{p['port']}/{p['proto']}" for p in a['open_ports'])
        row = [a['id'], a['ip'], a['mac'], a['vendor'], a['device_type'], a['device_model'],
               a['firmware_version'], a.get('serial',''), ports_str,
               ', '.join(a['protocols_ot']), a['zone'], a['criticality'],
               a['status'], a['notes'], '2026-03-19']
        ws1.append(row)
        if i % 2 == 1:
            for cell in ws1[ws1.max_row]:
                cell.fill = ALT_FILL
        for cell in ws1[ws1.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws1.row_dimensions[1].height = 30
    autofit(ws1)

    # ---- Sheet 2: Open Ports Detail ----
    ws2 = wb.create_sheet("Open Ports Detail")
    headers2 = ['IP','Porta','Protocollo','Servizio','Versione','Necessario','Note']
    ws2.append(headers2)
    style_header(ws2, 1)
    ws2.freeze_panes = 'A2'
    port_notes = {
        (172, 23): 'CRITICO: protocollo in chiaro',
        (445, ''): 'ATTENZIONE: SMB in rete OT',
        (50000,''): 'CRITICO: dati produzione in chiaro senza auth',
        (9600,''): 'ATTENZIONE: FINS senza auth',
        (515,''): 'ATTENZIONE: LPD obsoleto',
    }
    i = 0
    for a in ASSETS:
        for p in a['open_ports']:
            necessary = 'N' if p['service'] in ['Telnet','NetBIOS','MSRPC','LPD Printer'] else 'S'
            note = ''
            if p['service'] == 'Telnet': note = 'CRITICO: protocollo in chiaro'
            elif p['service'] == 'SMB': note = 'ATTENZIONE: SMB in rete OT'
            elif p['port'] == 50000: note = 'CRITICO: dati produzione esposti'
            elif p['service'] == 'FINS': note = 'ATTENZIONE: senza autenticazione'
            elif p['service'] == 'LPD Printer': note = 'ATTENZIONE: protocollo obsoleto'
            row = [a['ip'], p['port'], p['proto'], p['service'], p.get('version',''),
                   necessary, note]
            ws2.append(row)
            if i % 2 == 1:
                for cell in ws2[ws2.max_row]:
                    cell.fill = ALT_FILL
            for cell in ws2[ws2.max_row]:
                cell.border = border
            i += 1
    autofit(ws2)

    # ---- Sheet 3: Vulnerability Summary ----
    ws3 = wb.create_sheet("Vulnerability Summary")
    headers3 = ['Finding ID','IP Device','Modello','Titolo','Severita','CVSS Score',
                 'CVSS Vector','IEC 62443 SR','Remediation','Priorita']
    ws3.append(headers3)
    style_header(ws3, 1)
    ws3.freeze_panes = 'A2'
    for i, fin in enumerate(FINDINGS):
        row = [fin['id'], fin['device_ip'], fin['device_model'], fin['title'],
               fin['severity'], fin['cvss_score'], fin['cvss_vector'],
               '; '.join(fin['iec62443_sr']), fin['remediation'], fin['remediation_priority']]
        ws3.append(row)
        sev_col = ws3.cell(row=ws3.max_row, column=5)
        sev_color = SEVERITY_COLORS.get(fin['severity'], 'FFFFFF')
        sev_col.fill = PatternFill("solid", fgColor=sev_color)
        if fin['severity'] in ('Critical',):
            sev_col.font = Font(color='FFFFFF', bold=True)
        if i % 2 == 1:
            for j, cell in enumerate(ws3[ws3.max_row]):
                if j != 4:  # skip severity cell
                    cell.fill = ALT_FILL
        for cell in ws3[ws3.max_row]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    autofit(ws3)

    # ---- Sheet 4: IEC 62443 Compliance ----
    ws4 = wb.create_sheet("IEC 62443 Compliance")
    headers4 = ['SR Reference','Title','Status','Note']
    ws4.append(headers4)
    style_header(ws4, 1)
    ws4.freeze_panes = 'A2'
    STATUS_COLORS = {
        'Conforme': 'C6EFCE', 'Non Conforme': 'FFC7CE',
        'Parziale': 'FFEB9C', 'N/A': 'E0E0E0'
    }
    for i, sr in enumerate(SR_TABLE):
        ws4.append([sr['sr'], sr['title'], sr['status'], sr['note']])
        stat_cell = ws4.cell(row=ws4.max_row, column=3)
        stat_cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(sr['status'], 'FFFFFF'))
        if i % 2 == 1:
            for j, cell in enumerate(ws4[ws4.max_row]):
                if j != 2:
                    cell.fill = ALT_FILL
        for cell in ws4[ws4.max_row]:
            cell.border = border
    autofit(ws4)

    path = os.path.join(base_path, 'asset_inventory.xlsx')
    wb.save(path)
    print(f"asset_inventory.xlsx OK ({os.path.getsize(path)//1024}KB)")

# ============================================================
# GENERATE HTML REPORT
# ============================================================
def gen_html(base_path):
    sev_counts = {}
    for f in FINDINGS:
        sev_counts[f['severity']] = sev_counts.get(f['severity'], 0) + 1

    def sev_badge(sev):
        colors = {'Critical':'#c0392b','High':'#e67e22','Medium':'#f39c12','Low':'#27ae60','Info':'#2980b9'}
        bg = colors.get(sev, '#95a5a6')
        return f'<span style="background:{bg};color:white;padding:2px 8px;border-radius:3px;font-size:0.85em;font-weight:bold;">{sev}</span>'

    # SVG bar chart for executive summary
    max_count = max(sev_counts.values()) if sev_counts else 1
    svg_bars = ''
    sev_order = [('Critical','#c0392b'),('High','#e67e22'),('Medium','#f39c12'),('Low','#27ae60')]
    bar_w = 60
    for idx, (s, col) in enumerate(sev_order):
        cnt = sev_counts.get(s, 0)
        bh = int(cnt / max_count * 120) if cnt else 0
        x = 40 + idx * 90
        y = 140 - bh
        svg_bars += f'<rect x="{x}" y="{y}" width="{bar_w}" height="{bh}" fill="{col}" rx="3"/>'
        svg_bars += f'<text x="{x+bar_w//2}" y="{y-5}" text-anchor="middle" font-size="14" font-weight="bold" fill="{col}">{cnt}</text>'
        svg_bars += f'<text x="{x+bar_w//2}" y="158" text-anchor="middle" font-size="11" fill="#555">{s}</text>'

    # Zone/Conduit SVG diagram
    zone_svg = '''
    <svg viewBox="0 0 800 420" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:800px;border:1px solid #ddd;border-radius:8px;background:#fafafa">
      <!-- Zone 1: PLC Zone -->
      <rect x="20" y="40" width="220" height="340" rx="10" fill="#E3F2FD" stroke="#1565C0" stroke-width="2"/>
      <text x="130" y="65" text-anchor="middle" font-size="13" font-weight="bold" fill="#1565C0">Zone 1 - PLC Zone (SL-2)</text>
      <rect x="35" y="78" width="190" height="22" rx="4" fill="#1565C0"/><text x="130" y="93" text-anchor="middle" font-size="10" fill="white">Omron NJ501-1300 x5</text>
      <text x="130" y="110" text-anchor="middle" font-size="9" fill="#555">.21 .22 .23 .24 .111</text>
      <rect x="35" y="125" width="190" height="22" rx="4" fill="#1565C0"/><text x="130" y="140" text-anchor="middle" font-size="10" fill="white">B&amp;R Idecon Bilance x4</text>
      <text x="130" y="157" text-anchor="middle" font-size="9" fill="#555">.201 .202 .203 .204</text>
      <rect x="35" y="175" width="190" height="18" rx="3" fill="#ffcccc"/><text x="130" y="187" text-anchor="middle" font-size="9" fill="#c0392b">EtherNet/IP | FINS | OPC-UA | TCP/50000</text>
      <!-- Zone 2: HMI Zone -->
      <rect x="290" y="40" width="200" height="200" rx="10" fill="#E8F5E9" stroke="#2E7D32" stroke-width="2"/>
      <text x="390" y="65" text-anchor="middle" font-size="13" font-weight="bold" fill="#2E7D32">Zone 2 - HMI Zone (SL-1)</text>
      <rect x="305" y="78" width="170" height="22" rx="4" fill="#2E7D32"/><text x="390" y="93" text-anchor="middle" font-size="10" fill="white">Hakko GR-series x5</text>
      <text x="390" y="110" text-anchor="middle" font-size="9" fill="#555">.41 .42 .43 .44 .121</text>
      <rect x="305" y="125" width="170" height="18" rx="3" fill="#ffcccc"/><text x="390" y="137" text-anchor="middle" font-size="9" fill="#c0392b">HTTP (no TLS) | SSH (anomalo)</text>
      <!-- Zone 3: Remote/Infra Zone -->
      <rect x="290" y="265" width="200" height="120" rx="10" fill="#FFF3E0" stroke="#E65100" stroke-width="2"/>
      <text x="390" y="288" text-anchor="middle" font-size="12" font-weight="bold" fill="#E65100">Zone 3 - Remote Access (SL-2)</text>
      <rect x="305" y="298" width="170" height="20" rx="4" fill="#E65100"/><text x="390" y="312" text-anchor="middle" font-size="10" fill="white">HPE Aruba Switch .239.2</text>
      <rect x="305" y="325" width="170" height="20" rx="4" fill="#E65100"/><text x="390" y="339" text-anchor="middle" font-size="10" fill="white">Secomea SiteManager .239.254</text>
      <rect x="305" y="352" width="170" height="18" rx="3" fill="#ffcccc"/><text x="390" y="364" text-anchor="middle" font-size="9" fill="#c0392b">TELNET attivo! | HTTPS | SSH</text>
      <!-- Zone 4: Unclassified -->
      <rect x="545" y="40" width="230" height="120" rx="10" fill="#FAFAFA" stroke="#9E9E9E" stroke-width="2" stroke-dasharray="6,3"/>
      <text x="660" y="65" text-anchor="middle" font-size="12" font-weight="bold" fill="#9E9E9E">Zone 4 - Unclassified</text>
      <rect x="560" y="78" width="200" height="22" rx="4" fill="#9E9E9E"/><text x="660" y="93" text-anchor="middle" font-size="10" fill="white">Host Windows/VMware .224.10</text>
      <rect x="560" y="108" width="200" height="18" rx="3" fill="#ffaaaa"/><text x="660" y="120" text-anchor="middle" font-size="9" fill="#c0392b">SMB | MSRPC | VMware 902/912</text>
      <!-- Conduit C1: HMI -> PLC -->
      <line x1="290" y1="130" x2="240" y2="130" stroke="#2E7D32" stroke-width="2.5" stroke-dasharray="6,3" marker-end="url(#arr)"/>
      <line x1="240" y1="130" x2="240" y2="140" stroke="#2E7D32" stroke-width="2.5"/>
      <text x="263" y="125" text-anchor="middle" font-size="10" fill="#2E7D32" font-weight="bold">C1</text>
      <!-- Conduit C2: Remote -> PLC -->
      <line x1="290" y1="320" x2="255" y2="320" stroke="#E65100" stroke-width="2.5" stroke-dasharray="6,3"/>
      <line x1="255" y1="320" x2="255" y2="200" stroke="#E65100" stroke-width="2.5"/>
      <line x1="255" y1="200" x2="240" y2="200" stroke="#E65100" stroke-width="2.5" marker-end="url(#arr2)"/>
      <text x="258" y="270" text-anchor="middle" font-size="10" fill="#E65100" font-weight="bold">C2</text>
      <!-- C3: Unclassified -> PLC (danger) -->
      <line x1="545" y1="100" x2="490" y2="100" stroke="#c0392b" stroke-width="2.5"/>
      <line x1="490" y1="100" x2="490" y2="250" stroke="#c0392b" stroke-width="2.5"/>
      <line x1="490" y1="250" x2="240" y2="250" stroke="#c0392b" stroke-width="2.5" marker-end="url(#arr3)"/>
      <text x="380" y="243" text-anchor="middle" font-size="9" fill="#c0392b" font-weight="bold">C3 - NON CONTROLLATO</text>
      <!-- arrows -->
      <defs>
        <marker id="arr" markerWidth="8" markerHeight="6" refX="8" refY="3" orient="auto">
          <polygon points="0 0, 8 3, 0 6" fill="#2E7D32"/>
        </marker>
        <marker id="arr2" markerWidth="8" markerHeight="6" refX="8" refY="3" orient="auto">
          <polygon points="0 0, 8 3, 0 6" fill="#E65100"/>
        </marker>
        <marker id="arr3" markerWidth="8" markerHeight="6" refX="8" refY="3" orient="auto">
          <polygon points="0 0, 8 3, 0 6" fill="#c0392b"/>
        </marker>
      </defs>
    </svg>'''

    # Build findings HTML
    findings_html = ''
    for fin in FINDINGS:
        srs = ' '.join(f'<code style="background:#eef;padding:1px 4px;border-radius:2px">{s}</code>' for s in fin['iec62443_sr'])
        ev_html = fin['evidence'].replace('\n','<br>')
        rem_html = fin['remediation'].replace('\n','<br>')
        findings_html += f'''
        <div style="border:1px solid #ddd;border-left:4px solid {"#c0392b" if fin["severity"]=="Critical" else "#e67e22" if fin["severity"]=="High" else "#f39c12"};
             border-radius:6px;margin-bottom:20px;overflow:hidden;">
          <div style="padding:12px 16px;background:#f8f8f8;display:flex;justify-content:space-between;align-items:center;">
            <span style="font-weight:bold;font-size:1.05em;">{fin["id"]} — {fin["title"]}</span>
            {sev_badge(fin["severity"])}
          </div>
          <div style="padding:16px;">
            <table style="width:100%;border-collapse:collapse;font-size:0.9em;">
              <tr><td style="width:160px;color:#666;padding:4px 0;vertical-align:top"><b>Device:</b></td><td>{fin["device_ip"]} ({fin["device_model"]})</td></tr>
              <tr><td style="color:#666;padding:4px 0;vertical-align:top"><b>CVSS Score:</b></td><td><b>{fin["cvss_score"]}</b> &nbsp; <code style="font-size:0.85em;color:#555">{fin["cvss_vector"]}</code></td></tr>
              <tr><td style="color:#666;padding:4px 0;vertical-align:top"><b>IEC 62443 SR:</b></td><td>{srs}</td></tr>
              <tr><td style="color:#666;padding:4px 0;vertical-align:top"><b>Parte:</b></td><td>{fin["iec62443_part"]}</td></tr>
              <tr><td style="color:#666;padding:4px 0;vertical-align:top"><b>Priorita:</b></td><td><b>{fin["remediation_priority"]}</b></td></tr>
            </table>
            <div style="margin-top:12px"><b>Descrizione:</b><p style="margin:4px 0 0 0;line-height:1.6">{fin["description"]}</p></div>
            <details style="margin-top:12px">
              <summary style="cursor:pointer;color:#1565C0;font-weight:bold">Evidence (espandi)</summary>
              <pre style="background:#1e1e1e;color:#d4d4d4;padding:12px;border-radius:4px;overflow-x:auto;font-size:0.82em;margin-top:8px">{fin["evidence"]}</pre>
            </details>
            <div style="margin-top:12px;background:#f0f7f0;padding:12px;border-radius:4px;border-left:3px solid #27ae60">
              <b>Remediation:</b><br>{rem_html}
            </div>
          </div>
        </div>'''

    # Build assets HTML table
    assets_table_rows = ''
    for a in ASSETS:
        ports_str = '<br>'.join(f"{p['port']}/{p['proto']} {p['service']}" for p in a['open_ports'])
        crit_color = '#c0392b' if a['criticality']=='H' else '#e67e22' if a['criticality']=='M' else '#27ae60'
        note_style = 'color:#c0392b;font-weight:bold' if '***' in a.get('notes','') else ''
        assets_table_rows += f'''
        <tr>
          <td style="font-family:monospace">{a["id"]}</td>
          <td style="font-family:monospace">{a["ip"]}</td>
          <td style="font-size:0.85em">{a["vendor"]}</td>
          <td>{a["device_type"]}</td>
          <td>{a["device_model"]}</td>
          <td style="font-size:0.85em">{a["firmware_version"]}</td>
          <td style="font-size:0.8em">{ports_str}</td>
          <td style="font-size:0.85em">{", ".join(a["protocols_ot"]) or "—"}</td>
          <td style="font-size:0.85em">{a["zone"]}</td>
          <td style="text-align:center;color:{crit_color};font-weight:bold">{a["criticality"]}</td>
          <td style="font-size:0.8em;{note_style}">{a.get("notes","")}</td>
        </tr>'''

    # Build SR compliance table
    sr_rows = ''
    status_colors_bg = {'Conforme':'#c6efce','Non Conforme':'#ffc7ce','Parziale':'#ffeb9c','N/A':'#f0f0f0'}
    for sr in SR_TABLE:
        bg = status_colors_bg.get(sr['status'], 'white')
        sr_rows += f'''
        <tr>
          <td style="font-family:monospace;white-space:nowrap">{sr["sr"]}</td>
          <td>{sr["title"]}</td>
          <td style="text-align:center;background:{bg};font-weight:bold;white-space:nowrap">{sr["status"]}</td>
          <td style="font-size:0.85em;color:#555">{sr["note"]}</td>
        </tr>'''

    # Count compliance stats
    conf_count = sum(1 for s in SR_TABLE if s['status']=='Conforme')
    nc_count = sum(1 for s in SR_TABLE if s['status']=='Non Conforme')
    par_count = sum(1 for s in SR_TABLE if s['status']=='Parziale')
    na_count = sum(1 for s in SR_TABLE if s['status']=='N/A')

    html = f'''<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>OT Security Assessment Report — IEC 62443 — 2026-03-19</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, Helvetica, sans-serif; color: #333; margin: 0; background: #f5f6fa; }}
  .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
  /* Cover */
  .cover {{ background: linear-gradient(135deg, #1a237e 0%, #283593 50%, #1565C0 100%);
    color: white; padding: 60px 40px; border-radius: 10px; margin-bottom: 30px; text-align: center; }}
  .cover h1 {{ font-size: 2.2em; margin: 0 0 10px 0; }}
  .cover .subtitle {{ font-size: 1.2em; opacity: 0.85; margin-bottom: 30px; }}
  .cover .meta {{ display: flex; justify-content: center; gap: 40px; flex-wrap: wrap; margin-top: 20px; }}
  .cover .meta-item {{ background: rgba(255,255,255,0.1); padding: 10px 20px; border-radius: 6px; }}
  /* Sections */
  .section {{ background: white; border-radius: 8px; padding: 24px; margin-bottom: 20px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.08); }}
  .section h2 {{ color: #1565C0; border-bottom: 2px solid #1565C0; padding-bottom: 8px; margin-top: 0; }}
  .section h3 {{ color: #283593; margin-top: 20px; }}
  /* Tables */
  table {{ width: 100%; border-collapse: collapse; font-size: 0.9em; }}
  th {{ background: #1F3864; color: white; padding: 8px 10px; text-align: left; }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #eee; vertical-align: top; }}
  tr:nth-child(even) {{ background: #f8f9fc; }}
  tr:hover {{ background: #e8f0fe; }}
  /* Summary cards */
  .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 16px; margin: 20px 0; }}
  .card {{ padding: 16px; border-radius: 8px; text-align: center; color: white; }}
  .card .num {{ font-size: 2.5em; font-weight: bold; line-height: 1; }}
  .card .lbl {{ font-size: 0.9em; opacity: 0.9; margin-top: 4px; }}
  /* Print */
  @media print {{
    body {{ background: white; }}
    .container {{ padding: 0; }}
    .cover {{ border-radius: 0; }}
    .no-print {{ display: none !important; }}
    details {{ open: true; }}
    details summary {{ display: none; }}
    details[open] + * {{ display: block; }}
  }}
  /* Filter bar */
  .filter-bar {{ margin-bottom: 12px; }}
  .filter-bar input, .filter-bar select {{
    padding: 6px 10px; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px; font-size: 0.9em;
  }}
  code {{ background: #f0f0f0; padding: 1px 4px; border-radius: 3px; font-size: 0.9em; }}
  pre {{ white-space: pre-wrap; word-break: break-all; }}
</style>
</head>
<body>
<div class="container">

<!-- COVER -->
<div class="cover">
  <div style="font-size:3em;margin-bottom:10px">🏭</div>
  <h1>OT Security Assessment Report</h1>
  <div class="subtitle">Conformita IEC 62443 — Rete OT Industriale</div>
  <div class="meta">
    <div class="meta-item"><div style="font-size:0.8em;opacity:0.7">Assessment</div><div style="font-weight:bold">OT_Assessment_Impianto_X</div></div>
    <div class="meta-item"><div style="font-size:0.8em;opacity:0.7">Data</div><div style="font-weight:bold">2026-03-19</div></div>
    <div class="meta-item"><div style="font-size:0.8em;opacity:0.7">Assessor</div><div style="font-weight:bold">Nome Cognome</div></div>
    <div class="meta-item"><div style="font-size:0.8em;opacity:0.7">Subnet Analizzata</div><div style="font-weight:bold">172.16.224.0/20</div></div>
    <div class="meta-item"><div style="font-size:0.8em;opacity:0.7">Standard</div><div style="font-weight:bold">IEC 62443-3-3</div></div>
  </div>
  <div style="margin-top:30px">
    <button class="no-print" onclick="window.print()"
      style="background:rgba(255,255,255,0.2);color:white;border:1px solid rgba(255,255,255,0.5);
             padding:10px 24px;border-radius:6px;cursor:pointer;font-size:1em;">
      Stampa / Esporta PDF
    </button>
  </div>
</div>

<!-- EXECUTIVE SUMMARY -->
<div class="section">
  <h2>1. Executive Summary</h2>
  <p>Assessment di sicurezza OT/ICS su rete industriale in fase di commissioning (non in produzione).
  Il testing attivo e stato esplicitamente autorizzato dal responsabile dell'impianto.
  Scansione eseguita da Kali Linux VM (172.16.224.250) sulla subnet 172.16.224.0/20.</p>

  <div class="summary-grid">
    <div class="card" style="background:#1565C0"><div class="num">{len(ASSETS)}</div><div class="lbl">Asset Trovati</div></div>
    <div class="card" style="background:#c0392b"><div class="num">{sev_counts.get("Critical",0)}</div><div class="lbl">Critical</div></div>
    <div class="card" style="background:#e67e22"><div class="num">{sev_counts.get("High",0)}</div><div class="lbl">High</div></div>
    <div class="card" style="background:#f39c12"><div class="num">{sev_counts.get("Medium",0)}</div><div class="lbl">Medium</div></div>
    <div class="card" style="background:#27ae60"><div class="num">{sev_counts.get("Low",0)}</div><div class="lbl">Low</div></div>
    <div class="card" style="background:#8e44ad"><div class="num">3</div><div class="lbl">Asset Non Documentati</div></div>
  </div>

  <div style="display:flex;gap:24px;flex-wrap:wrap;align-items:flex-start;margin:20px 0">
    <div>
      <svg viewBox="0 0 450 175" xmlns="http://www.w3.org/2000/svg" style="width:450px;max-width:100%">
        <rect width="450" height="175" fill="#fafafa" rx="6" stroke="#ddd"/>
        <text x="225" y="20" text-anchor="middle" font-size="12" font-weight="bold" fill="#333">Finding per Severita</text>
        {svg_bars}
      </svg>
    </div>
    <div style="flex:1;min-width:250px">
      <h3 style="margin-top:0">Top 3 Raccomandazioni Immediate</h3>
      <ol style="padding-left:20px;line-height:2">
        <li><b>Disabilitare Telnet su HPE switch 172.16.239.2</b> — accesso admin in chiaro</li>
        <li><b>Bloccare accesso non autenticato TCP/50000 su bilance B&R</b> — dati produzione esposti</li>
        <li><b>Isolare/identificare host Windows .224.10</b> — SMB+VMware in zona OT</li>
      </ol>
      <h3>Gap di Conformita Principali IEC 62443</h3>
      <ul style="line-height:1.8;padding-left:20px">
        <li>SR 5.1/5.2: Nessuna segmentazione Zone/Conduit implementata</li>
        <li>SR 4.1: Protocolli in chiaro (Telnet, HTTP, TCP/50000)</li>
        <li>SR 2.1: Accesso non autenticato a FINS, EtherNet/IP, TCP/50000</li>
        <li>SR 7.8: Inventario incompleto (3 asset non documentati)</li>
      </ul>
    </div>
  </div>
</div>

<!-- ASSET INVENTORY -->
<div class="section">
  <h2>2. Asset Inventory</h2>
  <div class="filter-bar no-print">
    <input type="text" id="assetFilter" placeholder="Filtra per IP, vendor, modello..." onkeyup="filterTable('assetTable','assetFilter')" style="width:250px">
    <select onchange="filterByZone(this.value)">
      <option value="">Tutte le zone</option>
      <option value="Zone 1">Zone 1 - PLC</option>
      <option value="Zone 2">Zone 2 - HMI</option>
      <option value="Zone 3">Zone 3 - Remote</option>
      <option value="Zone 4">Zone 4 - Unclassified</option>
    </select>
  </div>
  <div style="overflow-x:auto">
  <table id="assetTable">
    <thead><tr>
      <th>ID</th><th>IP</th><th>Vendor</th><th>Tipo</th><th>Modello</th>
      <th>Firmware</th><th>Porte</th><th>Protocolli OT</th><th>Zone</th><th>Crit.</th><th>Note</th>
    </tr></thead>
    <tbody>{assets_table_rows}</tbody>
  </table>
  </div>
</div>

<!-- FINDINGS -->
<div class="section">
  <h2>3. Vulnerability Findings ({len(FINDINGS)} trovati)</h2>
  {findings_html}
</div>

<!-- ZONE & CONDUIT MAP -->
<div class="section">
  <h2>4. Zone &amp; Conduit Map (IEC 62443-3-2)</h2>
  <p>Proposta di segmentazione basata sui risultati della scansione. La rete attuale e <b>flat</b> — non sono state identificate separazioni fisiche o logiche tra le zone.</p>
  {zone_svg}
  <table style="margin-top:20px">
    <thead><tr><th>Conduit</th><th>From</th><th>To</th><th>Protocolli Ammessi</th><th>Stato Attuale</th></tr></thead>
    <tbody>
      <tr><td>C1</td><td>Zone 2 - HMI</td><td>Zone 1 - PLC</td><td>FINS, EtherNet/IP</td><td style="color:#e67e22">Non segregato — rete flat</td></tr>
      <tr><td>C2</td><td>Zone 3 - Remote</td><td>Zone 1 - PLC</td><td>HTTPS only</td><td style="color:#e67e22">Parziale — Telnet bypass</td></tr>
      <tr><td>C3</td><td>Zone 4 - Unclassified</td><td>Zone 1 - PLC</td><td><b>Nessuno — da bloccare</b></td><td style="color:#c0392b">Non controllato — SMB/VMware accede PLC zone</td></tr>
    </tbody>
  </table>
</div>

<!-- IEC 62443 COMPLIANCE -->
<div class="section">
  <h2>5. Mappatura IEC 62443-3-3 Security Requirements</h2>
  <div class="summary-grid">
    <div class="card" style="background:#27ae60"><div class="num">{conf_count}</div><div class="lbl">Conforme</div></div>
    <div class="card" style="background:#c0392b"><div class="num">{nc_count}</div><div class="lbl">Non Conforme</div></div>
    <div class="card" style="background:#f39c12"><div class="num">{par_count}</div><div class="lbl">Parziale</div></div>
    <div class="card" style="background:#95a5a6"><div class="num">{na_count}</div><div class="lbl">N/A</div></div>
  </div>
  <div style="overflow-x:auto">
  <table>
    <thead><tr><th style="width:90px">SR Ref.</th><th>Titolo</th><th style="width:120px">Stato</th><th>Note</th></tr></thead>
    <tbody>{sr_rows}</tbody>
  </table>
  </div>
</div>

<!-- REMEDIATION ROADMAP -->
<div class="section">
  <h2>6. Remediation Roadmap</h2>
  <h3 style="color:#c0392b">Immediate (entro 5 giorni lavorativi)</h3>
  <ul style="line-height:2">
    <li><b>[F-001]</b> Disabilitare Telnet su HPE Aruba switch (172.16.239.2)</li>
    <li><b>[F-002]</b> Implementare ACL switch per bloccare accesso non autorizzato TCP/50000</li>
    <li><b>[F-003]</b> Configurare IP Filter su tutti i PLC Omron NJ501-1300 per FINS</li>
    <li><b>[F-006]</b> Identificare e isolare host Windows/VMware 172.16.224.10</li>
  </ul>
  <h3 style="color:#e67e22">Short-term (entro 30 giorni)</h3>
  <ul style="line-height:2">
    <li><b>[F-004]</b> Implementare IP filtering su EtherNet/IP e valutare CIP Security</li>
    <li><b>[F-005]</b> Abilitare HTTPS o implementare reverse proxy TLS per HMI Hakko</li>
    <li><b>[F-007]</b> Documentare asset .111 e .121, investigare anomalie</li>
    <li><b>[F-008]</b> Verificare e configurare security mode OPC-UA (Sign&amp;Encrypt)</li>
    <li><b>[F-009]</b> Verificare e chiudere porta 8080 su Secomea se non necessaria</li>
    <li><b>[F-010]</b> Investigare e disabilitare SSH/LPD su HMI .121 se non necessari</li>
  </ul>
  <h3 style="color:#27ae60">Long-term (entro 90 giorni)</h3>
  <ul style="line-height:2">
    <li>Implementare segmentazione Zone/Conduit con firewall industriale (IEC 62443-3-2)</li>
    <li>Deploy sistema di monitoring continuo (SR 6.2) — es. Claroty, Dragos, Nozomi</li>
    <li>Definire processo di patch management per device OT</li>
    <li>Implementare network access control (NAC) per prevenire device non autorizzati</li>
    <li>Formazione personale su IEC 62443 e OT security awareness</li>
  </ul>
</div>

<!-- APPENDIX -->
<div class="section">
  <h2>7. Appendice — Comandi Eseguiti</h2>
  <details>
    <summary style="cursor:pointer;color:#1565C0;font-weight:bold">Mostra comandi nmap e tool utilizzati</summary>
    <pre style="background:#1e1e1e;color:#d4d4d4;padding:16px;border-radius:6px;overflow-x:auto;font-size:0.85em;margin-top:8px">
# FASE 1 - Setup
ssh kali@172.16.224.250 "hostname && ip a"

# FASE 2a - Host Discovery
sudo nmap -sn 172.16.224.0/20 -oX ~/assessment/raw/discovery.xml -oG ~/assessment/raw/discovery.txt
# Risultato: 18 host trovati (inclusi 2 non documentati: .111 e .121)

# FASE 2b - Port Scan (porte OT critiche)
sudo nmap -sS -p 21,22,23,25,80,102,443,502,515,4840,8080,9600,20000,44818,50000 --max-rate 50 -T2 --open [host] -oX ~/assessment/raw/portscan.xml

# FASE 2c - Service Version Detection
sudo nmap -sV --version-intensity 7 -p 80,443,4840,9600,44818,50000,8080,22,23,515 --max-rate 30 -T2 [host] -oX ~/assessment/raw/versions.xml

# FASE 2d - EtherNet/IP NSE Scan
sudo nmap -sU -p 44818 --script enip-info 172.16.224.21,22,23,24,111 -oX ~/assessment/raw/enip.xml

# FASE 2d - FINS Omron
sudo nmap -p 9600 --script omron-info 172.16.224.21,22,23,24,111 -oX ~/assessment/raw/fins.xml

# FASE 2e - HTTP Banners HMI Hakko
for ip in 172.16.224.41-44 172.16.239.2; do curl -s -v -m 5 http://$ip/; done

# FASE 2f - B&R TCP/50000 Banner Grab
for ip in 172.16.224.201-204; do echo "" | nc -v -w 3 $ip 50000; done
# FINDING CRITICO: dati produzione esposti in chiaro senza autenticazione

# FASE 2g - SNMP Enumeration Omron
for ip in .21-.24,.111; do snmpwalk -v2c -c public -t 3 $ip; done
# Risultato: timeout (SNMP community 'public' non accettata = buona pratica)

# FASE 2h - Unknown Device Probe (.10)
sudo nmap -Pn -sS -sU -p T:1-1024,U:161,162,67,68,69,514 172.16.224.10 -oX ~/assessment/raw/unknown_10.xml
# Identificato: Windows host con porte VMware (902, 912) e SMB
    </pre>
  </details>
</div>

<div style="text-align:center;color:#999;font-size:0.8em;padding:20px">
  Generato automaticamente | OT Security Assessment IEC 62443 | 2026-03-19 | Confidenziale
</div>

</div><!-- /container -->

<script>
function filterTable(tableId, filterId) {{
  var filter = document.getElementById(filterId).value.toUpperCase();
  var rows = document.getElementById(tableId).getElementsByTagName("tr");
  for (var i = 1; i < rows.length; i++) {{
    var found = rows[i].innerText.toUpperCase().indexOf(filter) > -1;
    rows[i].style.display = found ? "" : "none";
  }}
}}
function filterByZone(zone) {{
  var rows = document.getElementById("assetTable").getElementsByTagName("tr");
  for (var i = 1; i < rows.length; i++) {{
    if (!zone) {{ rows[i].style.display = ""; continue; }}
    var found = rows[i].innerText.indexOf(zone) > -1;
    rows[i].style.display = found ? "" : "none";
  }}
}}
</script>
</body>
</html>'''

    path = os.path.join(base_path, 'report_iec62443.html')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"report_iec62443.html OK ({os.path.getsize(path)//1024}KB)")

# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    BASE = '/home/kali/assessment/reports'
    os.makedirs(BASE, exist_ok=True)

    print("=== OT Assessment IEC 62443 - Report Generation ===")
    gen_csv(BASE)
    gen_excel(BASE)
    gen_html(BASE)

    # Save JSON data
    with open(os.path.join(BASE, 'assessment_data.json'), 'w', encoding='utf-8') as f:
        json.dump({"assessment":"OT_Assessment_Impianto_X","date":"2026-03-19",
                   "assessor":"Nome Cognome","assets":ASSETS,"findings":FINDINGS}, f,
                  ensure_ascii=False, indent=2)
    print("assessment_data.json OK")
    print(f"\nFiles in {BASE}:")
    for fn in os.listdir(BASE):
        sz = os.path.getsize(os.path.join(BASE, fn))
        print(f"  {fn}: {sz//1024}KB")
